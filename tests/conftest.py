import os
import time
from datetime import datetime
import pytest
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, TimeoutException
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# Global list to store test results
test_results = []
screenshots_dir = os.path.join(os.path.dirname(__file__), "screenshots")

@pytest.fixture(scope="session", autouse=True)
def create_directories():
    os.makedirs(screenshots_dir, exist_ok=True)

# Helper function with multiple selector fallbacks and stale element retry
def find_element_robust(driver, css_selector, xpath_selector=None, link_text=None, timeout=5):
    start_time = time.time()
    while time.time() - start_time < timeout:
        try:
            # 1. CSS Selector Try
            try:
                return WebDriverWait(driver, 1.5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, css_selector))
                )
            except (TimeoutException, StaleElementReferenceException):
                pass
                
            # 2. XPath Try
            if xpath_selector:
                try:
                    return WebDriverWait(driver, 1.5).until(
                        EC.presence_of_element_located((By.XPATH, xpath_selector))
                    )
                except (TimeoutException, StaleElementReferenceException):
                    pass
                    
            # 3. Link Text Try
            if link_text:
                try:
                    return WebDriverWait(driver, 1.5).until(
                        EC.presence_of_element_located((By.LINK_TEXT, link_text))
                    )
                except (TimeoutException, StaleElementReferenceException):
                    pass
            
            time.sleep(0.5)
        except StaleElementReferenceException:
            # Explicitly retry on stale elements
            time.sleep(0.5)
            continue
            
    raise TimeoutException(f"Element not found using fallbacks: CSS='{css_selector}', XPath='{xpath_selector}', Text='{link_text}'")

@pytest.fixture(scope="function")
def driver(request):
    app_url = os.getenv("APP_URL", "http://localhost:3000")
    
    options = Options()
    if os.getenv("HEADLESS", "true").lower() == "true":
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920,1080")
        
    options.add_argument("--ignore-certificate-errors")
    options.add_argument("--allow-insecure-localhost")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(0) # Explicit wait helper used instead of implicit
    
    # Attach helper to driver instance for easy access in tests
    driver.find_element_robust = lambda css, xpath=None, text=None, t=5: find_element_robust(driver, css, xpath, text, t)
    
    if request.node:
        request.node.funcargs['driver'] = driver
        
    driver.get(app_url)
    yield driver
    driver.quit()

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    
    if rep.when == "call":
        node_id = item.nodeid
        module = node_id.split("::")[0].replace("tests/", "").replace(".py", "").capitalize()
        test_case_name = node_id.split("::")[-1]
        
        duration = round(rep.duration, 3)
        status = rep.outcome.upper()
        error_msg = ""
        screenshot_path = ""
        
        if rep.failed:
            error_msg = str(rep.longreprtext)
            driver = item.funcargs.get("driver")
            if driver:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_name = "".join([c if c.isalnum() else "_" for c in test_case_name])
                screenshot_filename = f"{safe_name}_{timestamp}.png"
                screenshot_filepath = os.path.join(screenshots_dir, screenshot_filename)
                try:
                    driver.save_screenshot(screenshot_filepath)
                    screenshot_path = os.path.abspath(screenshot_filepath)
                except Exception as e:
                    error_msg += f"\n[Screenshot Capture Error]: {str(e)}"
                    
        test_results.append({
            "module": module,
            "name": test_case_name,
            "status": status,
            "duration": duration,
            "error": error_msg,
            "screenshot": screenshot_path,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    elif rep.when == "setup" and rep.outcome == "skipped":
        node_id = item.nodeid
        module = node_id.split("::")[0].replace("tests/", "").replace(".py", "").capitalize()
        test_case_name = node_id.split("::")[-1]
        test_results.append({
            "module": module,
            "name": test_case_name,
            "status": "SKIP",
            "duration": 0.0,
            "error": "Skipped",
            "screenshot": "",
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

def pytest_sessionfinish(session, exitstatus):
    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    total = len(test_results)
    passed = sum(1 for r in test_results if r["status"] == "PASS")
    failed = sum(1 for r in test_results if r["status"] == "FAIL")
    skipped = sum(1 for r in test_results if r["status"] == "SKIP")
    pass_rate = round((passed / total) * 100, 2) if total > 0 else 0.0
    
    ws_summary["A1"] = "Smart Ticketing E2E Test Run Summary"
    ws_summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    ws_summary["A3"] = "Metric"
    ws_summary["B3"] = "Value"
    ws_summary["A3"].fill = header_fill
    ws_summary["A3"].font = header_font
    ws_summary["B3"].fill = header_fill
    ws_summary["B3"].font = header_font
    
    metrics = [
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", failed),
        ("Skipped", skipped),
        ("Pass Rate %", f"{pass_rate}%"),
        ("Run Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for idx, (metric, val) in enumerate(metrics, start=4):
        ws_summary[f"A{idx}"] = metric
        ws_summary[f"B{idx}"] = val
        ws_summary[f"A{idx}"].font = Font(bold=True)
        ws_summary[f"B{idx}"].alignment = align_center
        
    ws_summary["D3"] = "Module Name"
    ws_summary["E3"] = "Total"
    ws_summary["F3"] = "Pass"
    ws_summary["G3"] = "Fail"
    ws_summary["H3"] = "Skip"
    for col in ["D", "E", "F", "G", "H"]:
        ws_summary[f"{col}3"].fill = header_fill
        ws_summary[f"{col}3"].font = header_font
        ws_summary[f"{col}3"].alignment = align_center
        
    module_data = {}
    for r in test_results:
        mod = r["module"]
        if mod not in module_data:
            module_data[mod] = {"total": 0, "pass": 0, "fail": 0, "skip": 0}
        module_data[mod]["total"] += 1
        if r["status"] == "PASS":
            module_data[mod]["pass"] += 1
        elif r["status"] == "FAIL":
            module_data[mod]["fail"] += 1
        elif r["status"] == "SKIP":
            module_data[mod]["skip"] += 1
            
    row_idx = 4
    for mod, counts in module_data.items():
        ws_summary[f"D{row_idx}"] = mod
        ws_summary[f"E{row_idx}"] = counts["total"]
        ws_summary[f"F{row_idx}"] = counts["pass"]
        ws_summary[f"G{row_idx}"] = counts["fail"]
        ws_summary[f"H{row_idx}"] = counts["skip"]
        for col in ["E", "F", "G", "H"]:
            ws_summary[f"{col}{row_idx}"].alignment = align_center
        row_idx += 1

    # 2. Test Details Sheet
    ws_details = wb.create_sheet(title="Test Details")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers_details = [
        "Module", "Test Case Name", "Status", "Duration (s)", 
        "Error Message", "Screenshot Path", "Timestamp"
    ]
    for col_idx, h in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_skip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    for r_idx, r in enumerate(test_results, start=2):
        ws_details.cell(row=r_idx, column=1, value=r["module"])
        ws_details.cell(row=r_idx, column=2, value=r["name"]).alignment = align_left
        
        status_cell = ws_details.cell(row=r_idx, column=3, value=r["status"])
        status_cell.alignment = align_center
        
        if r["status"] == "PASS":
            row_fill = fill_pass
        elif r["status"] == "FAIL":
            row_fill = fill_fail
        else:
            row_fill = fill_skip
            
        for c in range(1, 8):
            ws_details.cell(row=r_idx, column=c).fill = row_fill
            
        ws_details.cell(row=r_idx, column=4, value=r["duration"]).alignment = align_center
        ws_details.cell(row=r_idx, column=5, value=r["error"]).alignment = align_left
        ws_details.cell(row=r_idx, column=6, value=r["screenshot"]).alignment = align_left
        ws_details.cell(row=r_idx, column=7, value=r["timestamp"]).alignment = align_center
        
    for ws in [ws_summary, ws_details]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    report_path = os.path.join(os.path.dirname(__file__), "../test_report.xlsx")
    wb.save(report_path)
    print(f"\n[Test Report Saved]: {os.path.abspath(report_path)}")
